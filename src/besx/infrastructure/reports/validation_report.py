"""
validation_report.py

Módulo responsável por gerar relatórios detalhados de validação
com todos os cálculos intermediários exportados para Excel.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from besx.infrastructure.logging.logger import logger
import os

from typing import List, Dict, Any, Optional

def gerar_relatorio_validacao(file_manager: Any, config: Any, resultados_mensais: List[Dict[str, Any]], calculos_detalhados: List[Dict[str, Any]], prefixo: str = "") -> str:
    """
    Gera um relatório completo de validação em Excel com múltiplas sheets.
    
    Args:
        file_manager: Instância do FileManager para gerenciar caminhos
        config: Dicionário com toda a configuração da simulação
        resultados_mensais: Lista de dicionários com resultados mensais
        calculos_detalhados: Lista de dicionários com DataFrames detalhados por mês
        prefixo: Prefixo descritivo (opcional) para o nome final
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    logger.info("Gerando relatório de validação detalhado...")
    
    # Caminho do arquivo final
    nome_arquivo = f"relatorio_validacao_dimensionamento_{prefixo}.xlsx" if prefixo else "relatorio_validacao_dimensionamento.xlsx"
    caminho_excel = file_manager.get_data_path(nome_arquivo)
    
    # Cria o writer do Excel
    with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
        # Sheet 1: Configuração
        criar_sheet_configuracao(writer, config)
        
        # Sheet 2: Resumo Mensal
        df_resultados = pd.DataFrame(resultados_mensais)
        criar_sheet_resumo_mensal(writer, df_resultados)
        
        # Sheets 3+: Cálculos detalhados por mês
        criar_sheets_calculos_detalhados(writer, calculos_detalhados)
    
    # Aplica formatação
    formatar_workbook(caminho_excel)
    
    logger.info(f"Relatório de validação salvo em: {caminho_excel}")
    return caminho_excel


def criar_sheet_configuracao(writer: pd.ExcelWriter, config: Any) -> None:
    """
    Cria a sheet com todos os parâmetros de configuração.
    """
    dados_config = []
    
    # Seção: Simulação
    dados_config.append(['=== CONFIGURAÇÃO DA SIMULAÇÃO ===', ''])
    dados_config.append(['SOH Inicial (%)', config.simulacao.SOH_INICIAL_PERC])
    dados_config.append(['Anos de Simulação', config.simulacao.ANOS_SIMULACAO])
    dados_config.append(['Data Início', config.simulacao.data_inicio_simulacao])
    dados_config.append(['', ''])
    
    # Seção: Bateria
    dados_config.append(['=== PARÂMETROS DA BATERIA ===', ''])
    dados_config.append(['Capacidade Nominal (Wh)', config.bateria.capacidade_nominal_wh])
    dados_config.append(['Capacidade Limite de Perda (%)', config.bateria.capacidade_limite_perda_perc])
    dados_config.append(['Temperatura (K)', config.bateria.Tbat_kelvin])
    dados_config.append(['Resistência Interna (Ohms)', config.bateria.Rs])
    dados_config.append(['SOC Min', config.bateria.soc_min])
    dados_config.append(['SOC Max', config.bateria.soc_max])
    dados_config.append(['Potência BESS (W)', config.bateria.P_bess])
    dados_config.append(['Ah', config.bateria.Ah])
    dados_config.append(['Ns (células em série)', config.bateria.Ns])
    dados_config.append(['Np (células em paralelo)', config.bateria.Np])
    dados_config.append(['', ''])
    
    # Seção: Modelo de Degradação - Ciclo
    dados_config.append(['=== MODELO DE DEGRADAÇÃO - CICLO ===', ''])
    for key, value in config.modelo_degradacao.ciclo.model_dump().items():
        dados_config.append([f'ciclo.{key}', value])
    dados_config.append(['', ''])
    
    # Seção: Modelo de Degradação - Calendário
    dados_config.append(['=== MODELO DE DEGRADAÇÃO - CALENDÁRIO ===', ''])
    for key, value in config.modelo_degradacao.calendario.model_dump().items():
        dados_config.append([f'calendario.{key}', value])
    
    # Cria DataFrame e exporta
    df_config = pd.DataFrame(dados_config, columns=['Parâmetro', 'Valor'])
    df_config.to_excel(writer, sheet_name='Configuracao', index=False)


def criar_sheet_resumo_mensal(writer: pd.ExcelWriter, df_resultados: pd.DataFrame) -> None:
    """
    Cria a sheet com resumo mensal (similar ao resultados_completos.xlsx).
    """
    df_resultados.to_excel(writer, sheet_name='Resumo_Mensal', index=False)


def criar_sheets_calculos_detalhados(writer: pd.ExcelWriter, calculos_detalhados: List[Dict[str, Any]]) -> None:
    """
    Cria sheets detalhadas para cada mês (Ciclo e Calendário).
    
    Args:
        writer: ExcelWriter do pandas
        calculos_detalhados: Lista de dicts com 'mes', 'df_ciclo', 'df_calendario'
    """
    for calc in calculos_detalhados:
        mes_id = calc['mes']
        df_ciclo = calc['df_ciclo']
        df_calendario = calc['df_calendario']
        
        # Sheet de Degradação Cíclica
        nome_sheet_ciclo = f'Mes_{mes_id}_Ciclo'
        df_ciclo.to_excel(writer, sheet_name=nome_sheet_ciclo, index=False)
        
        # Sheet de Degradação Calendário
        nome_sheet_cal = f'Mes_{mes_id}_Calendario'
        df_calendario.to_excel(writer, sheet_name=nome_sheet_cal, index=False)


def formatar_workbook(caminho_excel: str) -> None:
    """
    Aplica formatação profissional ao workbook:
    - Cabeçalhos em negrito com cor de fundo
    - Ajusta largura das colunas
    - Congela painéis
    """
    wb = load_workbook(caminho_excel)
    
    # Cores
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formata cabeçalho (primeira linha)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limite de 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congela painel (primeira linha)
        ws.freeze_panes = 'A2'
    
    wb.save(caminho_excel)


def exportar_debug_degradacao(
    dados: Any,
    etapa: str,
    mes_id: Optional[int] = None,
    sufixo: str = "",
    pasta_debug: Optional[str] = None
) -> str:
    """
    Exporta dados intermediários do modelo de degradação para validação externa (Excel).
    """
    import os
    import pandas as pd
    import numpy as np

    # Se pasta_debug for fornecida, usa ela. Senão, cria local (fallback)
    if pasta_debug:
        pasta = pasta_debug
    else:
        pasta = "debug_degradacao"
        os.makedirs(pasta, exist_ok=True)

    if isinstance(dados, pd.DataFrame):
        df = dados.copy()
    elif isinstance(dados, pd.Series):
        df = dados.to_frame(name="valor")
    elif isinstance(dados, (list, np.ndarray)):
        if len(dados) > 0 and isinstance(dados[0], dict):
             df = pd.DataFrame(dados)
        elif len(dados) == 0:
             df = pd.DataFrame(columns=["valor"])
        else:
             df = pd.DataFrame(dados, columns=["valor"])
    else:
        raise TypeError("Tipo de dado não suportado para exportação.")

    partes_nome = [etapa]
    if mes_id is not None:
        partes_nome.append(f"mes_{mes_id}")
    if sufixo:
        partes_nome.append(sufixo)

    nome_arquivo = "_".join(partes_nome) + ".xlsx"
    caminho = os.path.join(pasta, nome_arquivo)
    
    df.to_excel(caminho, index=False)
    return caminho


def export_xlsx(df_list: List[pd.DataFrame], filename: str) -> None:
    """
    Exporta uma lista de DataFrames para um arquivo Excel com múltiplas abas.
    """
    logger.info(f"Salvando Excel em '{filename}'...")
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for i, df_para_salvar in enumerate(df_list):
                nome_planilha = f'Sheet_{i+1}'
                df_para_salvar.to_excel(writer, sheet_name=nome_planilha, index=False)
        logger.info(f"Arquivo '{filename}' salvo com sucesso.")
    except Exception as e:
        logger.error(f"Erro ao salvar Excel: {e}")
